import json
import subprocess
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook


SCRIPT = Path(__file__).with_name("batch_manifest_preflight.py")


def save_sheet(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Date", "Stock", "Model", "VIN"])
    sheet.append(["08/22/2026", "S2426", "2024 HYUNDAI ELANTRA", "KMHLL4DGXRU744841"])
    workbook.save(path)


def request_payload() -> dict:
    start = (datetime.now(timezone.utc) - timedelta(minutes=1)).isoformat()
    vins = ["WP1AA2A59KLB00525", "W1N0G8DBXNV374754", "4JGFB4KBXLA013794"]
    return {
        "request_id": "batch:3",
        "batch": {"vehicle_count": 3},
        "schedule": {"starts_at": start},
        "vehicles": [
            {
                "request_id": f"batch:3:{index}:vehicle",
                "vin": vin,
                "freight": {
                    "amount": 500,
                    "evidence": {
                        "loadPrice": 1500,
                        "distinctVinCount": 3,
                        "matchedRowNumbers": [index],
                    },
                },
            }
            for index, vin in enumerate(vins, start=1)
        ],
    }


def run(
    tmp_path: Path,
    payload: dict,
    sheet_setup=save_sheet,
) -> tuple[subprocess.CompletedProcess[str], dict]:
    request = tmp_path / "request.json"
    stock_sheet = tmp_path / "stock.xlsx"
    output = tmp_path / "result.json"
    request.write_text(json.dumps(payload), encoding="utf-8")
    sheet_setup(stock_sheet)
    completed = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--request",
            str(request),
            "--stock-sheet",
            str(stock_sheet),
            "--stock-prefix",
            "S",
            "--expected-request-id",
            "batch:3",
            "--output",
            str(output),
        ],
        text=True,
        capture_output=True,
    )
    return completed, json.loads(output.read_text(encoding="utf-8"))


def test_accepts_valid_batch_and_returns_sequential_candidates(tmp_path: Path) -> None:
    completed, result = run(tmp_path, request_payload())

    assert completed.returncode == 0
    assert result["ready"] is True
    assert [item["stock"] for item in result["sheet"]["candidates"]] == ["S2427", "S2428", "S2429"]
    assert all(item["check_digit_valid"] for item in result["checks"]["vin_validity"].values())


def test_rejects_duplicate_vin_or_bad_freight(tmp_path: Path) -> None:
    payload = request_payload()
    payload["vehicles"][1]["vin"] = payload["vehicles"][0]["vin"]
    payload["vehicles"][2]["freight"]["amount"] = 499

    completed, result = run(tmp_path, payload)

    assert completed.returncode == 2
    assert result["ready"] is False
    assert result["checks"]["unique_vins"] is False
    assert result["checks"]["freight"]["4JGFB4KBXLA013794"]["valid"] is False


def test_rejects_stale_request_identity(tmp_path: Path) -> None:
    payload = request_payload()
    payload["request_id"] = "batch:2"

    completed, result = run(tmp_path, payload)

    assert completed.returncode == 2
    assert result["ready"] is False
    assert result["checks"]["current_request_id_matches"] is False


def save_partial_tail_rows(path: Path) -> None:
    save_sheet(path)
    workbook = load_workbook(path)
    sheet = workbook["Sheet1"]
    rows = [
        [None, "S2427", None, "WP1AA2A59KLB00525", 2019, "PORSCHE", "MACAN", "SILVER", 66180],
        [None, "S2428", None, "W1N0G8DBXNV374754", 2022, "MERCEDES-BENZ", "GLC 300", "GRAY", 77950],
        [None, "S2429", None, "4JGFB4KBXLA013794", 2020, "MERCEDES-BENZ", "GLE 350 4MATIC", "SILVER", 73218],
    ]
    for values in rows:
        sheet.append(values)
    workbook.save(path)


def test_resumes_exact_consecutive_payload_rows_at_sheet_tail(tmp_path: Path) -> None:
    completed, result = run(tmp_path, request_payload(), save_partial_tail_rows)

    assert completed.returncode == 0
    assert result["ready"] is True
    assert result["sheet"]["plan_mode"] == "RESUME_EXISTING_TAIL_ROWS"
    assert [item["stock"] for item in result["sheet"]["candidates"]] == ["S2427", "S2428", "S2429"]
    assert [item["row"] for item in result["sheet"]["candidates"]] == [3, 4, 5]


def save_mixed_rows(path: Path) -> None:
    save_sheet(path)
    workbook = load_workbook(path)
    sheet = workbook["Sheet1"]
    sheet.append([None, "S2427", None, "WP1AA2A59KLB00525"])
    workbook.save(path)


def test_reuses_unique_existing_rows_and_appends_only_missing_vins(tmp_path: Path) -> None:
    completed, result = run(tmp_path, request_payload(), save_mixed_rows)

    assert completed.returncode == 0
    assert result["ready"] is True
    assert result["sheet"]["plan_mode"] == "REUSE_EXISTING_AND_APPEND"
    assert [item["stock"] for item in result["sheet"]["candidates"]] == [
        "S2427",
        "S2428",
        "S2429",
    ]
    assert [item["sheet_action"] for item in result["sheet"]["candidates"]] == [
        "REUSE_EXISTING",
        "APPEND_NEW",
        "APPEND_NEW",
    ]


def save_duplicate_existing_vin(path: Path) -> None:
    save_mixed_rows(path)
    workbook = load_workbook(path)
    sheet = workbook["Sheet1"]
    sheet.append([None, "S2428", None, "WP1AA2A59KLB00525"])
    workbook.save(path)


def test_rejects_duplicate_existing_payload_vin(tmp_path: Path) -> None:
    completed, result = run(tmp_path, request_payload(), save_duplicate_existing_vin)

    assert completed.returncode == 2
    assert result["ready"] is False
    assert result["sheet"]["plan_mode"] == "MIXED_EXISTING_ROWS"


def save_non_tail_rows(path: Path) -> None:
    save_partial_tail_rows(path)
    workbook = load_workbook(path)
    sheet = workbook["Sheet1"]
    sheet.append([None, "S2430", None, "JTDBR32E720045761"])
    workbook.save(path)


def test_reuses_unique_payload_rows_that_are_not_the_sheet_tail(tmp_path: Path) -> None:
    completed, result = run(tmp_path, request_payload(), save_non_tail_rows)

    assert completed.returncode == 0
    assert result["ready"] is True
    assert result["sheet"]["plan_mode"] == "REUSE_EXISTING_ROWS"
    assert all(
        item["sheet_action"] == "REUSE_EXISTING"
        for item in result["sheet"]["candidates"]
    )
