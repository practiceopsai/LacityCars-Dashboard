#!/usr/bin/env python3
"""Validate a dashboard batch and reserve collision-free sheet candidates.

This helper is deliberately read-only. It replaces ad-hoc validation code in
live Hermes runs with one deterministic gate for schedule, VINs, freight,
duplicate sheet rows, and the next sequential store stock numbers.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


VIN_RE = re.compile(r"^[A-HJ-NPR-Z0-9]{17}$")
VIN_TRANSLITERATION = {
    **{str(number): number for number in range(10)},
    "A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7, "H": 8,
    "J": 1, "K": 2, "L": 3, "M": 4, "N": 5, "P": 7, "R": 9,
    "S": 2, "T": 3, "U": 4, "V": 5, "W": 6, "X": 7, "Y": 8, "Z": 9,
}
VIN_WEIGHTS = [8, 7, 6, 5, 4, 3, 2, 10, 0, 9, 8, 7, 6, 5, 4, 3, 2]


def normalize_vin(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def expected_check_digit(vin: str) -> str:
    remainder = sum(
        VIN_TRANSLITERATION[character] * weight
        for character, weight in zip(vin, VIN_WEIGHTS)
    ) % 11
    return "X" if remainder == 10 else str(remainder)


def as_number(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(str(value).replace("$", "").replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def freight_check(vehicle: dict[str, Any]) -> dict[str, Any]:
    freight = vehicle.get("freight") or {}
    evidence = freight.get("evidence") or {}
    supplied = as_number(freight.get("amount"))
    load_price = as_number(evidence.get("loadPrice"))
    distinct_vins = as_number(evidence.get("distinctVinCount"))
    matched_rows = evidence.get("matchedRowNumbers") or []
    calculated = load_price / distinct_vins if load_price is not None and distinct_vins else None
    valid = (
        supplied is not None
        and calculated is not None
        and abs(supplied - calculated) < 0.005
        and bool(matched_rows)
    )
    return {
        "supplied": supplied,
        "calculated": calculated,
        "matched_row_count": len(matched_rows),
        "valid": valid,
    }


def parse_start(value: object) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def scan_stock_sheet(path: Path, prefix: str, vins: list[str]) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    stock_pattern = re.compile(rf"^{re.escape(prefix.upper())}(\d+)$")
    stock_rows: list[tuple[int, int, str, str]] = []
    vin_counts = {vin: 0 for vin in vins}
    for row_number, values in enumerate(rows, start=1):
        normalized_values = [normalize_vin(value) for value in values]
        row_vins = set(vins).intersection(normalized_values)
        for vin in row_vins:
            vin_counts[vin] += 1
        row_vin = next(iter(row_vins), "")
        for value in values:
            stock = str(value or "").strip().upper()
            match = stock_pattern.fullmatch(stock)
            if match:
                stock_rows.append((int(match.group(1)), row_number, stock, row_vin))
                break

    if not stock_rows:
        raise ValueError(f"No {prefix.upper()} stock numbers were found in {path}")

    stock_rows.sort()
    max_number, max_row, max_stock, max_vin = stock_rows[-1]
    width = max(4, len(str(max_number)))
    existing_numbers = {number for number, _, _, _ in stock_rows}
    candidates = []
    for offset, vin in enumerate(vins, start=1):
        number = max_number + offset
        row_number = max_row + offset
        values_before = list(rows[row_number - 1][:11]) if row_number <= len(rows) else [None] * 11
        values_before += [None] * (11 - len(values_before))
        candidates.append(
            {
                "vin": vin,
                "stock": f"{prefix.upper()}{number:0{width}d}",
                "row": row_number,
                "row_values_before": values_before,
                "stock_collision_count": int(number in existing_numbers),
            }
        )

    return {
        "path": str(path.resolve()),
        "worksheet": worksheet.title,
        "last_sequential_stock": {"stock": max_stock, "row": max_row, "vin": max_vin},
        "payload_vin_counts": vin_counts,
        "candidates": candidates,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--request", type=Path, required=True)
    parser.add_argument("--stock-sheet", type=Path, required=True)
    parser.add_argument("--stock-prefix", required=True)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    errors: list[str] = []
    if not args.request.is_file():
        errors.append(f"request not found: {args.request}")
    if not args.stock_sheet.is_file():
        errors.append(f"stock sheet not found: {args.stock_sheet}")

    request: dict[str, Any] = {}
    vehicles: list[dict[str, Any]] = []
    if not errors:
        request = json.loads(args.request.read_text(encoding="utf-8"))
        vehicles = request.get("vehicles") or []

    vins = [normalize_vin(vehicle.get("vin")) for vehicle in vehicles]
    request_ids = [str(vehicle.get("request_id") or "") for vehicle in vehicles]
    expected_count = int((request.get("batch") or {}).get("vehicle_count") or len(vehicles))
    start = parse_start((request.get("schedule") or {}).get("starts_at"))
    now = datetime.now(timezone.utc)

    vin_checks: dict[str, Any] = {}
    for vin in vins:
        format_valid = bool(VIN_RE.fullmatch(vin))
        expected = expected_check_digit(vin) if format_valid else None
        vin_checks[vin] = {
            "format_valid": format_valid,
            "expected_check_digit": expected,
            "actual_check_digit": vin[8] if len(vin) == 17 else None,
            "check_digit_valid": bool(format_valid and expected == vin[8]),
        }

    checks = {
        "count_matches": len(vehicles) == expected_count and len(vehicles) > 0,
        "unique_request_ids": len(request_ids) == len(set(request_ids)) and all(request_ids),
        "complete_request_ids": all(":" in request_id for request_id in request_ids),
        "unique_vins": len(vins) == len(set(vins)) and all(vins),
        "vin_validity": vin_checks,
        "freight": {vin: freight_check(vehicle) for vin, vehicle in zip(vins, vehicles)},
        "authorized": bool(start and now >= start),
        "scheduled_start": start.isoformat() if start else None,
    }

    sheet: dict[str, Any] = {}
    if not errors:
        try:
            sheet = scan_stock_sheet(args.stock_sheet, args.stock_prefix, vins)
        except Exception as exc:  # provide a compact fail-closed artifact
            errors.append(str(exc))

    ready = not errors and all(
        [
            checks["count_matches"],
            checks["unique_request_ids"],
            checks["complete_request_ids"],
            checks["unique_vins"],
            checks["authorized"],
            all(item["format_valid"] and item["check_digit_valid"] for item in vin_checks.values()),
            all(item["valid"] for item in checks["freight"].values()),
            all(count == 0 for count in sheet.get("payload_vin_counts", {}).values()),
            all(
                candidate["stock_collision_count"] == 0
                and all(value is None for value in candidate["row_values_before"])
                for candidate in sheet.get("candidates", [])
            ),
        ]
    )
    result = {
        "ready": ready,
        "generated_at_utc": now.isoformat(),
        "checks": checks,
        "sheet": sheet,
        "errors": errors,
    }
    args.output.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({
        "ready": ready,
        "output": str(args.output.resolve()),
        "stocks": [candidate["stock"] for candidate in sheet.get("candidates", [])],
        "errors": errors,
    }))
    return 0 if ready else 2


if __name__ == "__main__":
    sys.exit(main())
